import os
import re
import sqlite3
import json
import requests
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, Response
from samba_utils import (
    discover_acl_resources,
    get_samba_users,
    get_user_groups,
    add_user_to_group,
    remove_user_from_group,
    create_samba_user,
    block_samba_user,
    unblock_samba_user,
    reset_samba_password,
    rename_samba_user,
    reset_user_samba_sessions
)

# Загрузка конфигурации
CONFIG_PATH = os.getenv("SAMBA_WEB_CONFIG", "config.json")
config = {
    "mode": "node",
    "secret_key": "f2a8c3d9b4e578c1d2e3f4a5b6c7d8e9",
    "central_auth_url": "http://host.docker.internal:5001",
    "node_api_token": "vpn-panel-shared-secret-token-2026",
    "bind_host": "0.0.0.0",
    "bind_port": 5002,
    "smb_conf_path": "/etc/samba/smb.conf",
    "verify_ssl": True
}

# 1. Загружаем из файла конфигурации, если он есть
if os.path.exists(CONFIG_PATH):
    try:
        with open(CONFIG_PATH, "r") as f:
            file_config = json.load(f)
            if "bind_port" in file_config:
                file_config["bind_port"] = int(file_config["bind_port"])
            if "verify_ssl" in file_config:
                if isinstance(file_config["verify_ssl"], str):
                    file_config["verify_ssl"] = file_config["verify_ssl"].lower() == "true"
            config.update(file_config)
    except Exception as e:
        print(f"Error loading config: {e}")

# 2. Переменные окружения имеют высший приоритет и переопределяют значения из файла
if os.getenv("SECRET_KEY"):
    config["secret_key"] = os.getenv("SECRET_KEY")
if os.getenv("CENTRAL_AUTH_URL"):
    config["central_auth_url"] = os.getenv("CENTRAL_AUTH_URL")
if os.getenv("NODE_API_TOKEN"):
    config["node_api_token"] = os.getenv("NODE_API_TOKEN")
if os.getenv("BIND_HOST"):
    config["bind_host"] = os.getenv("BIND_HOST")
if os.getenv("BIND_PORT"):
    config["bind_port"] = int(os.getenv("BIND_PORT"))
if os.getenv("SMB_CONF_PATH"):
    config["smb_conf_path"] = os.getenv("SMB_CONF_PATH")
if os.getenv("VERIFY_SSL"):
    config["verify_ssl"] = os.getenv("VERIFY_SSL").lower() == "true"

app = Flask(__name__)
app.secret_key = config["secret_key"]

# База данных аудита
NODE_DB = os.getenv("DATABASE_PATH", "samba-node.db")

def init_db():
    conn = sqlite3.connect(NODE_DB)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            username TEXT NOT NULL,
            action TEXT NOT NULL,
            details TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS share_links (
            token TEXT PRIMARY KEY,
            expires_at TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            description TEXT
        )
    """)
    conn.commit()
    conn.close()

def log_action(username, action, details=""):
    try:
        conn = sqlite3.connect(NODE_DB)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO audit_logs (username, action, details) VALUES (?, ?, ?)",
            (username, action, details)
        )
        conn.commit()
        # Очистка старых логов (храним только последние 1000)
        cursor.execute("""
            DELETE FROM audit_logs 
            WHERE id NOT IN (
                SELECT id FROM audit_logs 
                ORDER BY id DESC 
                LIMIT 1000
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error logging action: {e}")

# Декоратор для проверки авторизации
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            # Если запрос является API-запросом, возвращаем 401
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated_function

@app.context_processor
def inject_hostname():
    import socket
    return {"hostname": socket.gethostname()}

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if session.get("logged_in"):
        return redirect(url_for("index"))
        
    if request.method == 'POST':
        data = request.json or {}
        username = data.get("username", "").strip()
        password = data.get("password", "")
        
        if not username or not password:
            return jsonify({"error": "Укажите имя пользователя и пароль"}), 400
            
        # Запрос к Central Auth Server
        central_url = f"{config['central_auth_url'].rstrip('/')}/api/auth/verify"
        try:
            verify_ssl = config.get("verify_ssl", True)
            if not verify_ssl:
                import urllib3
                urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

            resp = requests.post(
                central_url,
                json={"username": username, "password": password},
                headers={"X-Node-Token": config["node_api_token"]},
                timeout=5,
                verify=verify_ssl
            )
            if resp.status_code == 200:
                res_data = resp.json()
                if res_data.get("success"):
                    session["logged_in"] = True
                    session["username"] = username
                    log_action(username, "LOGIN", "Успешный вход в систему")
                    return jsonify({"success": True})
                else:
                    return jsonify({"error": res_data.get("error", "Неверные учетные данные")}), 401
            elif resp.status_code == 403:
                return jsonify({"error": "Ошибка авторизации ноды на сервере авторизации"}), 403
            else:
                return jsonify({"error": f"Ошибка сервера авторизации (код {resp.status_code})"}), 500
        except requests.exceptions.RequestException as e:
            return jsonify({"error": f"Не удалось связаться с сервером авторизации: {e}"}), 500
            
    return render_template("login.html")

@app.route('/logout')
def logout():
    username = session.get("username", "unknown")
    log_action(username, "LOGOUT", "Выход из системы")
    session.pop("logged_in", None)
    session.pop("username", None)
    return redirect(url_for("login_page"))

@app.route('/favicon.ico')
def favicon():
    svg_icon = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="#c52f31"><path d="M19 12h-2v3h-3v2h3v3h3v-3h2v-2h-2v-3zm-2-6H5c-1.11 0-2 .9-2 2v12c0 1.11.9 2 2 2h8v-2h-8V8h12v3h2V8c0-1.11-.9-2-2-2z"/></svg>'
    return Response(svg_icon, mimetype='image/svg+xml')

@app.route('/')
@login_required
def index():
    return render_template("index.html")

# API эндпоинты
@app.route('/api/status')
@login_required
def api_status():
    return jsonify({
        "username": session.get("username"),
        "smb_conf_path": config["smb_conf_path"]
    })

@app.route('/api/matrix')
@login_required
def api_matrix():
    shares = discover_acl_resources(config["smb_conf_path"])
    users = get_samba_users()
    
    matrix = {}
    for user in users:
        username = user["username"]
        user_groups = get_user_groups(username)
        
        matrix[username] = {}
        for share in shares:
            rw_grp = share["rw_group"]
            ro_grp = share["ro_group"]
            
            # Проверяем права по группам
            if rw_grp and rw_grp in user_groups:
                matrix[username][share["name"]] = "rw"
            elif ro_grp and ro_grp in user_groups:
                matrix[username][share["name"]] = "ro"
            else:
                matrix[username][share["name"]] = "none"
                
    return jsonify({
        "shares": shares,
        "users": users,
        "matrix": matrix
    })

def generate_excel_matrix(shares, users):
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Сортируем пользователей по имени
    users.sort(key=lambda u: (u.get("full_name") or u["username"]).lower())
    active_users = [u for u in users if not u.get("disabled")]

    user_groups = {u["username"]: get_user_groups(u["username"]) for u in active_users}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Матрица прав"

    # Стили
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_cell = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_title = PatternFill(start_color="C52F31", end_color="C52F31", fill_type="solid")
    fill_header = PatternFill(start_color="121216", end_color="121216", fill_type="solid")
    fill_rw = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid") # Зеленый (есть доступ)
    fill_ro = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Желтый (только чтение)
    fill_unrestricted = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Заголовок
    ws.merge_cells("A1:C1")
    ws["A1"] = "Матрица прав доступа Samba"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40

    ws.row_dimensions[2].height = 15

    # Заголовки столбцов
    ws["A3"] = "Общая папка / Путь"
    ws["A3"].font = font_header
    ws["A3"].fill = fill_header
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A3"].border = thin_border
    
    ws.row_dimensions[3].height = 30

    col_idx = 2
    for user in active_users:
        col_letter = get_column_letter(col_idx)
        cell_ref = f"{col_letter}3"
        display_name = f"{user.get('full_name') or '—'}\n({user['username']})"
        ws[cell_ref] = display_name
        ws[cell_ref].font = font_header
        ws[cell_ref].fill = fill_header
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[cell_ref].border = thin_border
        col_idx += 1

    # Заполнение строками
    row_idx = 4
    for share in shares:
        ws.row_dimensions[row_idx].height = 25
        
        indent = "    " * share.get("depth", 0)
        disp_name = share.get("display_name", share["name"])
        share_desc = f"{indent}{disp_name}\n{share['path']}"
        ws.cell(row=row_idx, column=1, value=share_desc)
        ws.cell(row=row_idx, column=1).font = font_bold
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=1).border = thin_border

        is_configured = share["rw_group"] or share["ro_group"]
        
        col_idx = 2
        for user in active_users:
            username = user["username"]
            groups = user_groups[username]
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = font_cell
            
            if not is_configured:
                cell.value = "Без ограничений"
                cell.fill = fill_unrestricted
            else:
                rw_grp = share["rw_group"]
                ro_grp = share["ro_group"]
                
                if rw_grp and rw_grp in groups:
                    cell.value = "RW"
                    cell.fill = fill_rw
                elif ro_grp and ro_grp in groups:
                    cell.value = "RO"
                    cell.fill = fill_ro
                else:
                    cell.value = "—"
            
            col_idx += 1
        
        row_idx += 1

    # Авто-ширина
    max_len = 30
    for row in range(3, row_idx):
        val = ws.cell(row=row, column=1).value
        if val:
            lines = str(val).split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
    ws.column_dimensions['A'].width = min(max_len + 4, 50)

    for col in range(2, col_idx):
        col_letter = get_column_letter(col)
        max_len = 12
        val = ws.cell(row=3, column=col).value
        if val:
            lines = str(val).split('\n')
            max_len = max(len(l) for l in lines)
        ws.column_dimensions[col_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

@app.route('/api/matrix/export')
@login_required
def api_matrix_export():
    shares = discover_acl_resources(config["smb_conf_path"])
    users = get_samba_users()
    try:
        excel_data = generate_excel_matrix(shares, users)
        filename = f"Samba_Access_Matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/shared/matrix/<token>/export')
def api_shared_matrix_export(token):
    try:
        conn = sqlite3.connect(NODE_DB)
        cursor = conn.cursor()
        cursor.execute("SELECT expires_at FROM share_links WHERE token = ?", (token,))
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return "Ссылка не найдена или отозвана", 404
            
        expires_at = row[0]
        if expires_at < datetime.now().isoformat():
            return "Срок действия ссылки истек", 403
            
        shares = discover_acl_resources(config["smb_conf_path"])
        users = get_samba_users()
        
        excel_data = generate_excel_matrix(shares, users)
        filename = f"Samba_Access_Matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return f"Системная ошибка: {e}", 500

@app.route('/api/permissions/update', methods=['POST'])
@login_required
def api_permissions_update():
    data = request.json or {}
    username = data.get("username")
    share_name = data.get("share_name")
    access = data.get("access") # 'rw', 'ro', 'none'
    
    if not username or not share_name or access not in ['rw', 'ro', 'none']:
        return jsonify({"error": "Неверные параметры запроса"}), 400
        
    shares = discover_acl_resources(config["smb_conf_path"])
    # Находим настройки папки
    share = next((s for s in shares if s["name"] == share_name), None)
    if not share:
        return jsonify({"error": f"Каталог '{share_name}' не найден"}), 404
        
    rw_grp = share["rw_group"]
    ro_grp = share["ro_group"]
    
    if not rw_grp and not ro_grp:
        return jsonify({"error": f"Для каталога '{share_name}' не заданы группы доступа в ACL"}), 400
        
    success = True
    errors = []
    
    # Сначала удаляем из обеих групп, чтобы избежать конфликтов
    if rw_grp:
        remove_user_from_group(username, rw_grp)
    if ro_grp:
        remove_user_from_group(username, ro_grp)
        
    # Добавляем в нужную группу
    if access == 'rw':
        if rw_grp:
            if not add_user_to_group(username, rw_grp):
                success = False
                errors.append(f"Не удалось добавить в группу {rw_grp}")
        else:
            success = False
            errors.append("Группа чтения-записи не настроена для этой папки")
    elif access == 'ro':
        if ro_grp:
            if not add_user_to_group(username, ro_grp):
                success = False
                errors.append(f"Не удалось добавить в группу {ro_grp}")
        else:
            # Если группы RO нет, но есть RW, то права RW дают и чтение, и запись.
            # Но если пользователь просит только чтение, а группы RO нет, выдаем ошибку.
            success = False
            errors.append("Группа только для чтения не настроена для этой папки")
            
    if success:
        log_action(
            session["username"],
            "UPDATE_PERMISSIONS",
            f"Изменены права пользователя {username} на шару {share_name} -> {access.upper()}"
        )
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "errors": errors}), 500

@app.route('/api/users/create', methods=['POST'])
@login_required
def api_users_create():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    full_name = data.get("full_name", "").strip()
    
    if not username or not password or not full_name:
        return jsonify({"error": "Логин, пароль и ФИО сотрудника обязательны"}), 400
        
    # Защита от спецсимволов в имени пользователя (логине)
    if not re.match(r"^[a-zA-Z0-9_.-]+$", username):
        return jsonify({"error": "Недопустимые символы в логине пользователя"}), 400
        
    success, msg = create_samba_user(username, password, full_name)
    if success:
        log_action(session["username"], "CREATE_USER", f"Создан пользователь Samba: {username} ({full_name})")
        return jsonify({"success": True, "message": msg})
    else:
        return jsonify({"error": msg}), 500

@app.route('/api/users/reset_password', methods=['POST'])
@login_required
def api_users_reset_password():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    
    if not username or not password:
        return jsonify({"error": "Имя пользователя и новый пароль обязательны"}), 400
        
    success, msg = reset_samba_password(username, password)
    if success:
        log_action(session["username"], "RESET_PASSWORD", f"Сброшен пароль пользователя Samba: {username}")
        return jsonify({"success": True, "message": msg})
    else:
        return jsonify({"error": msg}), 500

@app.route('/api/users/toggle_block', methods=['POST'])
@login_required
def api_users_toggle_block():
    data = request.json or {}
    username = data.get("username")
    block = data.get("block") # boolean
    
    if not username or block is None:
        return jsonify({"error": "Неверные параметры запроса"}), 400
        
    if block:
        success = block_samba_user(username)
        action_text = "Блокировка"
    else:
        success = unblock_samba_user(username)
        action_text = "Разблокировка"
        
    if success:
        log_action(
            session["username"],
            "TOGGLE_USER_BLOCK",
            f"{'Заблокирован' if block else 'Разблокирован'} пользователь Samba: {username}"
        )
        return jsonify({"success": True})
    else:
        return jsonify({"error": f"Не удалось выполнить действие: {action_text}"}), 500

@app.route('/api/users/reset-sessions', methods=['POST'])
@login_required
def api_users_reset_sessions():
    data = request.json or {}
    username = data.get("username", "").strip()
    
    if not username:
        return jsonify({"error": "Имя пользователя обязательно"}), 400
        
    success, msg = reset_user_samba_sessions(username)
    if success:
        log_action(
            session["username"],
            "RESET_SESSIONS",
            f"Сброшены активные сессии пользователя Samba: {username} ({msg})"
        )
        return jsonify({"success": True, "message": msg})
    else:
        return jsonify({"error": msg}), 500

@app.route('/api/audit_logs')
@login_required
def api_audit_logs():
    try:
        conn = sqlite3.connect(NODE_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT timestamp, username, action, details FROM audit_logs ORDER BY id DESC LIMIT 150")
        rows = cursor.fetchall()
        conn.close()
        
        logs = []
        for r in rows:
            logs.append({
                "timestamp": r["timestamp"],
                "username": r["username"],
                "action": r["action"],
                "details": r["details"]
            })
        return jsonify(logs)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/share_links/create', methods=['POST'])
@login_required
def api_share_links_create():
    import secrets
    from datetime import datetime, timedelta
    
    data = request.json or {}
    description = data.get("description", "").strip()
    days = int(data.get("days", 7))
    
    token = secrets.token_hex(16)
    expires_at = (datetime.now() + timedelta(days=days)).isoformat()
    
    try:
        conn = sqlite3.connect(NODE_DB)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO share_links (token, expires_at, description) VALUES (?, ?, ?)",
            (token, expires_at, description)
        )
        conn.commit()
        conn.close()
        
        share_url = f"{request.scheme}://{request.host}/shared/matrix/{token}"
        
        log_action(
            session["username"],
            "CREATE_SHARE_LINK",
            f"Создана временная ссылка для просмотра: {description} (истекает {expires_at})"
        )
        
        return jsonify({"success": True, "token": token, "share_url": share_url})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/share_links', methods=['GET'])
@login_required
def api_share_links():
    try:
        conn = sqlite3.connect(NODE_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT token, expires_at, created_at, description FROM share_links ORDER BY created_at DESC")
        rows = cursor.fetchall()
        conn.close()
        
        links = []
        for r in rows:
            share_url = f"{request.scheme}://{request.host}/shared/matrix/{r['token']}"
            links.append({
                "token": r["token"],
                "expires_at": r["expires_at"],
                "created_at": r["created_at"],
                "description": r["description"],
                "share_url": share_url
            })
        return jsonify(links)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/share_links/revoke', methods=['POST'])
@login_required
def api_share_links_revoke():
    data = request.json or {}
    token = data.get("token")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
        
    try:
        conn = sqlite3.connect(NODE_DB)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM share_links WHERE token = ?", (token,))
        conn.commit()
        conn.close()
        
        log_action(session["username"], "REVOKE_SHARE_LINK", f"Отозвана временная ссылка: {token}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/shared/matrix/<token>')
def shared_matrix_page(token):
    from datetime import datetime
    try:
        conn = sqlite3.connect(NODE_DB)
        cursor = conn.cursor()
        cursor.execute("SELECT expires_at, description FROM share_links WHERE token = ?", (token,))
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return "Ссылка не найдена или была отозвана", 404
            
        expires_at, description = row
        if expires_at < datetime.now().isoformat():
            return "Срок действия ссылки истек", 403
            
        return render_template("shared_matrix.html", token=token, description=description, expires_at=expires_at)
    except Exception as e:
        return f"Системная ошибка: {e}", 500

@app.route('/api/shared/matrix/<token>')
def api_shared_matrix_data(token):
    from datetime import datetime
    try:
        conn = sqlite3.connect(NODE_DB)
        cursor = conn.cursor()
        cursor.execute("SELECT expires_at FROM share_links WHERE token = ?", (token,))
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return jsonify({"error": "Ссылка не найдена или отозвана"}), 404
            
        expires_at = row[0]
        if expires_at < datetime.now().isoformat():
            return jsonify({"error": "Срок действия ссылки истек"}), 403
            
        shares = discover_acl_resources(config["smb_conf_path"])
        users = get_samba_users()
        
        matrix = {}
        for user in users:
            username = user["username"]
            user_groups = get_user_groups(username)
            
            matrix[username] = {}
            for share in shares:
                rw_grp = share["rw_group"]
                ro_grp = share["ro_group"]
                
                if rw_grp and rw_grp in user_groups:
                    matrix[username][share["name"]] = "rw"
                elif ro_grp and ro_grp in user_groups:
                    matrix[username][share["name"]] = "ro"
                else:
                    matrix[username][share["name"]] = "none"
                    
        return jsonify({
            "shares": shares,
            "users": users,
            "matrix": matrix
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/users/edit', methods=['POST'])
@login_required
def api_users_edit():
    data = request.json or {}
    old_username = data.get("old_username", "").strip()
    new_username = data.get("new_username", "").strip()
    new_fullname = data.get("new_fullname", "").strip()
    
    if not old_username or not new_username or not new_fullname:
        return jsonify({"error": "Все поля обязательны"}), 400
        
    # Валидация логина
    if not re.match(r"^[a-zA-Z0-9_.-]+$", new_username):
        return jsonify({"error": "Недопустимые символы в логине пользователя"}), 400
        
    success, msg = rename_samba_user(old_username, new_username, new_fullname)
    if success:
        log_action(
            session["username"],
            "EDIT_USER",
            f"Изменен пользователь: {old_username} -> {new_username} ({new_fullname})"
        )
        return jsonify({"success": True, "message": msg})
    else:
        return jsonify({"error": msg}), 500


if __name__ == '__main__':
    init_db()
    app.run(host=config["bind_host"], port=config["bind_port"])
